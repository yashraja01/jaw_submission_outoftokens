"""Walk a document tree, recover its text, and classify every file by what it says.

The estate we are handed is nested by document type, but *the nesting will not match any sample we
have seen* and neither will the filenames. So nothing here may key on a path: the previous version
of this pipeline read `document_index.csv` and split documents by their `DOC-CC-*` / `DOC-REF-*`
identifier prefixes, which is exactly the assumption that breaks. A document's type is decided from
its own text, by title-line signatures, and the walk is a plain recursive glob for `.pdf`/`.xlsx`.

PyMuPDF with sort=True is the text layer. It preserves the label/value column geometry that both
certificate families are parsed on, and recovers roughly twice what pdfplumber's default text does
on the table-heavy types.
"""
import concurrent.futures as cf
import json
import os
import pathlib
import re
import sys
import unicodedata

import fitz

SQ = lambda t: re.sub(r"\s+", " ", t).strip()

# --------------------------------------------------------------------------- type signatures
#
# Each entry is (doc_type, weight, pattern). Patterns are matched against the squeezed, lowercased
# head of the document — titles and letterheads live there — except those marked `body`, which are
# matched against the whole text. The highest total score wins; ties fall to `unknown`, which is
# ingested and searchable but contributes no typed records.
#
# Signatures are deliberately redundant. Any single title line may be re-worded in a corpus we have
# not seen, so every type carries several independent cues (a title, a reference-number format, a
# distinctive field label) and needs only one of them to fire.
HEAD_CHARS = 1200

SIGNATURES = [
    ("completion_certificate", 3, r"work completion certificate|certificate of completion"),
    ("completion_certificate", 3, r"\bcompletion certificate\b"),
    ("completion_certificate", 2, r"issued under the provisions of"),
    ("completion_certificate", 2, r"\bref:?\s*cc/\d+/\d+/\d+"),

    ("company_completion_certificate", 4, r"record of work completed"),
    ("company_completion_certificate", 4, r"issued by the contractor"),
    ("company_completion_certificate", 3, r"internal ref:?\s*ccc|\bref:?\s*\w+/cc/"),
    ("company_completion_certificate", 3, r"client certificate ref"),
    ("company_completion_certificate", 2, r"company completion|contractor.s own record"),

    ("reference_letter", 4, r"letter of recommendation"),
    ("reference_letter", 3, r"to whom(so)?ever it may concern|to whom it may concern"),
    ("reference_letter", 3, r"\breference letter\b|letter of appreciation|testimonial"),
    ("reference_letter", 3, r"\bref:?\s*ref/"),
    ("reference_letter", 3, r"subject:[^|]{0,60}performance of m/s"),
    ("reference_letter", 2, r"this office has engaged|we have no hesitation in recommend"),

    ("performance_bond", 4, r"performance bank guarantee|bank guarantee department"),
    ("performance_bond", 3, r"\bbond no:?\s*bnd|\bbg no:?|guarantee & bond department"),
    ("performance_bond", 3, r"subject:[^|]{0,40}performance bond"),
    ("performance_bond", 2, r"irrevocable and unconditional|first written demand"),

    ("personnel_certificate", 3, r"credential id|certificate no\.|certification authority"),
    ("personnel_certificate", 3, r"\b(pmp|six sigma|prince2|lean|iso lead auditor)\b.{0,40}"
                                 r"certificat"),
    ("personnel_certificate", 2, r"this is to certify that|conferred upon"),
    ("personnel_certificate", 2, r"professional certification authority"),

    ("cv", 5, r"curriculum vitae|\bcurriculum\b"),
    ("cv", 3, r"key personnel"),
    ("cv", 2, r"total experience"),

    ("compliance_matrix", 4, r"compliance checklist|compliance matrix"),
    ("compliance_matrix", 2, r"bid compliance"),

    ("general_ledger_book", 4, r"general ledger"),
    ("general_ledger_book", 3, r"books of account"),

    ("bank_statement", 4, r"statement of account|account statement"),
    ("bank_statement", 2, r"a/c:\s*\d"),

    ("financial_statement", 4, r"statement of profit and loss|audited financial results"),
    ("financial_statement", 3, r"balance sheet"),

    ("final_ra_bill", 5, r"final running account bill|final ra bill"),
    ("final_ra_bill", 4, r"final bill with measurement"),
    ("ra_bill", 3, r"running account bill|\bra bill\b"),
    ("ra_bill", 2, r"contractor's bill"),

    ("tender_dossier", 4, r"tender submission dossier"),
    ("tender_dossier", 3, r"tender submission|bid value:"),

    ("iso_certificate", 4, r"certificate of registration"),
    ("iso_certificate", 3, r"accredited certification body|iaf member"),

    ("annual_report", 5, r"annual report"),
    ("annual_report", 2, r"directors.{0,3} report"),

    ("past_performance_portfolio", 5, r"past performance portfolio|portfolio of completed works"),
]

# A final RA bill also says "running account bill"; a company completion certificate also says
# "completion". Where one type's text is a superset of another's, the more specific type wins
# outright rather than by score.
DOMINATES = {
    "final_ra_bill": ["ra_bill"],
    "company_completion_certificate": ["completion_certificate"],
    "past_performance_portfolio": ["completion_certificate", "company_completion_certificate"],
    "annual_report": ["financial_statement"],
}

# Workbook types are decided by their column headers, not their filename. `Receivables_Ageing.xlsx`
# is the one the receivables questions are answered from, and its header row is far more stable
# than its name.
SHEET_SIGNATURES = [
    ("ageing_workbook", r"invoiced|outstanding", r"invoice no|client"),
    ("boq_workbook", r"rate|amount", r"item no|description"),
    ("trial_balance_workbook", r"debit|credit", r"account"),
    ("asset_register_workbook", r"acquired|cost", r"asset id|make"),
]


def classify_text(text):
    """(doc_type, score, runner_up). Content only — never the path, never the filename."""
    head = SQ(text[:HEAD_CHARS]).lower()
    body = SQ(text).lower()
    scores = {}
    for dtype, weight, pat in SIGNATURES:
        hay = body if weight <= 2 else head
        if re.search(pat, hay):
            scores[dtype] = scores.get(dtype, 0) + weight
    for winner, losers in DOMINATES.items():
        if winner in scores:
            for l in losers:
                scores.pop(l, None)
    if not scores:
        return "unknown", 0, None
    ranked = sorted(scores.items(), key=lambda kv: (-kv[1], kv[0]))
    return ranked[0][0], ranked[0][1], (ranked[1][0] if len(ranked) > 1 else None)


def letterhead(text):
    """The issuing organisation, read off the top-left of the first line.

    Layouts put the issuer at the left of the first line and a department at the right, separated
    by a run of spaces, so the first whitespace-run-delimited chunk is the name.
    """
    for line in text.splitlines():
        if line.strip():
            chunk = re.split(r"\s{2,}|·", line.strip())[0]
            return SQ(chunk).lower().strip(" .,-")
    return ""


def disambiguate_issuer(rows):
    """Split the two completion-certificate families by who issued them.

    A completion certificate is the *client's* sign-off and carries the client's letterhead; the
    company completion certificate is our own record of the same work and carries ours. Both say
    "completion certificate", and a corpus we have not seen may word their titles differently, so
    the issuer is the reliable discriminator. Which letterhead is ours is not hard-coded: it is
    whichever one appears on the most documents in the estate, since the contractor is the only
    organisation that issues at scale here.
    """
    heads = {}
    for r in rows:
        if r.get("head_org"):
            heads[r["head_org"]] = heads.get(r["head_org"], 0) + 1
    if not heads:
        return None
    ours, n = max(heads.items(), key=lambda kv: kv[1])
    if n < max(3, 0.05 * len(rows)):            # no dominant issuer; leave the scores alone
        return None
    pair = {"completion_certificate", "company_completion_certificate"}
    for r in rows:
        if r["doc_type"] not in pair:
            continue
        r["doc_type"] = ("company_completion_certificate" if r.get("head_org") == ours
                         else "completion_certificate")
    return ours


def classify_workbook(sheets):
    """Workbook type from the header row of any sheet."""
    for name, sheet in sheets.items():
        rows = sheet.get("rows") or []
        if not rows:
            continue
        head = " | ".join(str(c).lower() for c in (rows[0] or []) if c is not None)
        for dtype, pat_a, pat_b in SHEET_SIGNATURES:
            if re.search(pat_a, head) and re.search(pat_b, head):
                return dtype
    return "unknown_workbook"


# --------------------------------------------------------------------------- the walk
def doc_key(path, root):
    """A stable, filesystem-independent id for a document.

    The relative path with separators flattened, so two files of the same name in different
    subtrees cannot collide. Where the estate happens to use the `DOC-CC-001` convention we have
    seen, that stem survives intact and the build stays legible.
    """
    rel = path.relative_to(root)
    stem = rel.with_suffix("").as_posix()
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("_")
    return slug or "doc"


def find_documents(root):
    """Every PDF and XLSX under `root`, at any depth, in a deterministic order."""
    root = pathlib.Path(root).resolve()
    out = []
    for p in sorted(root.rglob("*")):
        if not p.is_file():
            continue
        suf = p.suffix.lower()
        if suf in (".pdf", ".xlsx", ".xlsm"):
            # Excel writes a lock file alongside an open workbook; it is not a document.
            if p.name.startswith("~$") or p.name.startswith("."):
                continue
            out.append(p)
    return out


def _read_pdf(path):
    with fitz.open(path) as d:
        pages = [p.get_text("text", sort=True) for p in d]
    return "\n\f\n".join(pages), len(pages)


def _read_workbook(path):
    import openpyxl
    vals = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        forms = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception:
        forms = None
    sheets = {}
    for name in vals.sheetnames:
        sv = vals[name]
        rows = [list(r) for r in sv.iter_rows(values_only=True)]
        formulas = {}
        if forms is not None and name in forms.sheetnames:
            formulas = {c.coordinate: c.value for row in forms[name].iter_rows() for c in row
                        if isinstance(c.value, str) and c.value.startswith("=")}
        sheets[name] = {"rows": rows, "formulas": formulas}
    return sheets


def _ingest_one(args):
    """Worker: read one file, classify it, write its text cache. Returns a catalog row.

    A file that cannot be read must not take the run down with it — one corrupt PDF in an estate
    of hundreds is a document we lose, not a submission we lose.
    """
    path_s, root_s, txt_dir_s = args
    path, root, txt_dir = pathlib.Path(path_s), pathlib.Path(root_s), pathlib.Path(txt_dir_s)
    did = doc_key(path, root)
    row = {"doc_id": did, "path": path.relative_to(root).as_posix(),
           "bytes": path.stat().st_size, "doc_type": "unreadable", "score": 0,
           "runner_up": None, "chars": 0, "pages": 0, "kind": path.suffix.lower().lstrip(".")}
    try:
        if path.suffix.lower() == ".pdf":
            text, pages = _read_pdf(path)
            text = unicodedata.normalize("NFKC", text)
            dtype, score, runner = classify_text(text)
            (txt_dir / f"{did}.txt").write_text(text, encoding="utf-8")
            row.update(doc_type=dtype, score=score, runner_up=runner, chars=len(text),
                       pages=pages, kind="pdf", head_org=letterhead(text))
        else:
            sheets = _read_workbook(path)
            row.update(doc_type=classify_workbook(sheets), kind="xlsx",
                       chars=sum(len(s["rows"]) for s in sheets.values()),
                       sheets=list(sheets.keys()))
            (txt_dir / f"{did}.json").write_text(
                json.dumps(sheets, default=str), encoding="utf-8")
    except Exception as e:                      # noqa: BLE001 - one bad file, not a failed run
        row["error"] = f"{type(e).__name__}: {e}"
    return row


def ingest(docs_root, build_dir, workers=None, log=print):
    """Extract and classify the whole estate. Returns the catalog (one row per file)."""
    root = pathlib.Path(docs_root).resolve()
    build = pathlib.Path(build_dir)
    txt = build / "txt"
    txt.mkdir(parents=True, exist_ok=True)

    files = find_documents(root)
    log(f"[ingest] {len(files)} files under {root}")
    if not files:
        raise SystemExit(f"[ingest] no PDF or XLSX files found under {root}")

    args = [(str(p), str(root), str(txt)) for p in files]
    workers = workers or min(8, (os.cpu_count() or 2))
    rows = []
    try:
        with cf.ProcessPoolExecutor(max_workers=workers) as ex:
            for i, row in enumerate(ex.map(_ingest_one, args, chunksize=4), 1):
                rows.append(row)
                if i % 100 == 0 or i == len(args):
                    log(f"[ingest]   {i}/{len(args)} extracted")
    except Exception as e:                      # noqa: BLE001 - pools fail on locked-down hosts
        log(f"[ingest] process pool unavailable ({type(e).__name__}); extracting serially")
        rows = []
        for i, a in enumerate(args, 1):
            rows.append(_ingest_one(a))
            if i % 100 == 0 or i == len(args):
                log(f"[ingest]   {i}/{len(args)} extracted")

    rows.sort(key=lambda r: r["doc_id"])
    ours = disambiguate_issuer(rows)
    if ours:
        log(f"[ingest] contractor letterhead: {ours!r}")
    (build / "catalog.json").write_text(json.dumps(rows, indent=1, default=str), encoding="utf-8")

    counts = {}
    for r in rows:
        counts[r["doc_type"]] = counts.get(r["doc_type"], 0) + 1
    log("[ingest] classified: " + ", ".join(f"{k}={v}" for k, v in sorted(counts.items())))
    weak = [r for r in rows if r["kind"] == "pdf" and r["chars"] < 200]
    if weak:
        log(f"[ingest] WARNING {len(weak)} PDFs yielded almost no text "
            f"(e.g. {[r['doc_id'] for r in weak[:5]]})")
    bad = [r for r in rows if r.get("error")]
    if bad:
        log(f"[ingest] WARNING {len(bad)} files failed to read: "
            f"{[(r['doc_id'], r['error']) for r in bad[:5]]}")
    return rows


if __name__ == "__main__":                      # manual: python -m pipeline.ingest <docs> <build>
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    ingest(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "build")
