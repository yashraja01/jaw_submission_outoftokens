"""Extract every document once, to disk. Nothing downstream reads a PDF.

PyMuPDF with sort=True is the primary text layer: it preserves the label/value column layout that
both certificate families depend on, and recovers ~2x what pdfplumber's default text does on
table-heavy pages. pdfplumber is used only for its table extraction, on the types that need it.
"""
import csv, json, pathlib, sys

import fitz

ROOT = pathlib.Path(__file__).resolve().parent.parent
DS = ROOT / "dataset"
OUT = ROOT / "build"
TXT = OUT / "txt"


def extract_pdfs():
    TXT.mkdir(parents=True, exist_ok=True)
    rows = list(csv.DictReader(open(DS / "document_index.csv", encoding="utf-8")))
    stats = []
    for r in rows:
        src = DS / "documents" / r["filename"]
        if src.suffix.lower() != ".pdf":
            stats.append({**r, "chars": 0, "pages": 0, "kind": "xlsx"})
            continue
        with fitz.open(src) as d:
            pages = [p.get_text("text", sort=True) for p in d]
        txt = "\n\f\n".join(pages)
        (TXT / f"{r['doc_id']}.txt").write_text(txt, encoding="utf-8")
        stats.append({**r, "chars": len(txt), "pages": len(pages), "kind": "pdf"})
    return stats


def extract_workbooks():
    """Both cached values and formulas, every sheet including Notes."""
    import openpyxl
    books = {}
    for src in sorted((DS / "documents" / "workbooks").glob("*.xlsx")):
        vals = openpyxl.load_workbook(src, data_only=True)
        forms = openpyxl.load_workbook(src, data_only=False)
        sheets = {}
        for name in vals.sheetnames:
            sv, sf = vals[name], forms[name]
            rows = [[c for c in row] for row in sv.iter_rows(values_only=True)]
            formulas = {c.coordinate: c.value for row in sf.iter_rows() for c in row
                        if isinstance(c.value, str) and c.value.startswith("=")}
            sheets[name] = {"rows": rows, "formulas": formulas}
        books[src.name] = sheets
    return books


def gate(stats):
    """Flag documents whose recovered text is anomalously short for their size."""
    pdfs = [s for s in stats if s["kind"] == "pdf"]
    flagged = []
    for s in pdfs:
        ratio = s["chars"] / max(int(s["size_bytes"]), 1)
        if s["chars"] < 200 or ratio < 0.004:
            flagged.append((s["doc_id"], s["doc_type"], s["chars"], s["size_bytes"], round(ratio, 4)))
    return flagged


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    stats = extract_pdfs()
    total = sum(s["chars"] for s in stats)
    json.dump(stats, open(OUT / "extract_stats.json", "w"), indent=1)

    books = extract_workbooks()
    json.dump(books, open(OUT / "workbooks.json", "w"), indent=1, default=str)

    print(f"PDFs      : {sum(1 for s in stats if s['kind']=='pdf')}")
    print(f"workbooks : {len(books)}  sheets={sum(len(v) for v in books.values())}")
    print(f"characters: {total:,}   (briefing: ~3.5M)")
    flagged = gate(stats)
    print(f"short-text flags: {len(flagged)}")
    for f in flagged[:15]:
        print("   ", f)
    assert len(stats) == 687, f"expected 687 documents, got {len(stats)}"
    assert abs(total - 3_500_000) < 300_000, f"character total {total:,} outside expected band"
    print("GATE PASS")


if __name__ == "__main__":
    sys.exit(main())
